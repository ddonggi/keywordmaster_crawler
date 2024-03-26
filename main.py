import openpyxl
import requests
from bs4 import BeautifulSoup
import requests
import time
import random

# 엑셀 파일 불러오기
fileName = input("엑셀 파일명을 입력하세요:")
print('filename:', fileName)
wb = openpyxl.load_workbook(fileName + '.xlsx')

if fileName.__contains__("_result"):
    fileName = fileName.replace("_result", "")
    print("기존에 진행중이던 파일 입니다")
    # print('filename:', fileName)

ws = wb.active

# F 컬럼에서 값이 있는 마지막 행 찾기

# 엑셀파일 첫 행 초기화
# 키워드 /모바일검색량	/pc 검색량	/문서 수	/총조회수	/비율
ws.cell(row=1, column=1, value='키워드')
ws.cell(row=1, column=2, value='모바일 검색량')
ws.cell(row=1, column=3, value='pc 검색량')
ws.cell(row=1, column=4, value='문서 수')
ws.cell(row=1, column=5, value='총 조회수')
ws.cell(row=1, column=6, value='비율')

count = 0
for row in ws:
    if not all([cell.value is None for cell in row]):
        count += 1
print(f'총 row : {count}')

last_row = 2  # 두 번째 행부터 시작
for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):
    if row[5] is not None:  # F 컬럼 값이 있는지 확인
        last_row += 1
    else:
        break

print('시작 행:', last_row)


def request_keyword(payload):
    try:
        # POST 요청 보내기
        response = requests.post(url, data=payload)

        # 응답 확인
        if response.status_code == 200:
            print("검색 요청이 성공하였습니다.")
            print("응답 데이터>>")

            if len(response.text) == 0:
                print('응답 데이터 내용이 없습니다. 20분간 대기후 재진행 합니다')
                time.sleep(1200)  # 20분간 대기
                request_keyword(payload)

            # JSON 응답 파싱
            data = response.json()
            # print(f'data:{data}')

            # 모바일 검색량
            mobile_search_volume = data['mo']
            print(f'  모바일 검색량 | {mobile_search_volume}')
            ws.cell(row=row_index, column=2, value=mobile_search_volume)

            # pc 검색량
            pc_search_volume = data['pc']
            print(f'  pc 검색량 | {pc_search_volume}')
            ws.cell(row=row_index, column=3, value=pc_search_volume)

            # 문서 수
            post = data['post']
            print(f'  문서 수 | {post}')
            ws.cell(row=row_index, column=4, value=post)

            # 총 조회수
            sum = data['sum']
            print(f'  총 조회수 | {sum}')
            ws.cell(row=row_index, column=5, value=sum)

            # 비율
            ratio = data['byul']
            ratio = ratio.replace("<span style=\"font-size:12px;color:#666;\">", "")
            ratio = ratio.replace("</span>", "")
            print(f'  비율 | {ratio}')
            ws.cell(row=row_index, column=6, value=ratio)

            # 이 코드는 검색 및 결과를 추가하는 예시입니다.
            # search_result = f'검색결과_{word}'  # 임시로 검색 결과 생성
            # search_result = f'검색결과_test'  # 임시로 검색 결과 생성
            # 결과를 기존 엑셀 파일의 F 컬럼에 추가

            # 10단어 마다 중간 저장
            if (row_index - last_row) % 10 == 0 and (row_index - last_row) > 1:
                wb.save(fileName + '_result.xlsx')
                print(f'{row_index - last_row}개의 단어를 검색하여 result 파일에 중간 저장되었습니다.')

            # 300단어 마다 중간 저장
            if row_index % 300 == 0:
                wb.save(fileName + str(row_index) + '.xlsx')
                print(f'{row_index}만큼의 검색량을 중간 저장 하였습니다.')
        else:
            print(f"오류 발생! 응답 코드: {response.status_code}")

    except Exception as e:
        print("오류 발생:", e)
        print("오류 발생:", e.with_traceback())


# 검색 결과를 기존 엑셀 파일의 B,C,D,E,F 컬럼에 추가
for row_index, row in enumerate(ws.iter_rows(min_row=last_row, max_col=1, values_only=True), start=last_row):
    search_query = row[0]  # A 컬럼에 있는 단어 (=검색어)
    random_sec = random.uniform(2, 5)
    print("-------------------------------------")
    print(f"랜덤 대기 시간: {random_sec} 초")
    time.sleep(random_sec)
    print("-------------------------------------")
    print('검색어 : "' + search_query + '"')
    # 단어를 검색하고 결과를 가져오는 로직
    # 요청 URL
    url = 'https://whereispost.com/keyword/functionmase.php'
    # form data 설정
    payload = {
        'query': search_query,
        's': 'true',
        'queries': ''
    }
    request_keyword(payload)


# 최종 결과 저장
wb.save(fileName + '_result.xlsx')
print(f'최종 검색 결과가 저장 되었습니다.')

# 엑셀 파일 닫기
wb.close()